import os
import re
import sys
import getpass
import threading
from PyQt6 import QtWidgets, QtCore, QtGui
from PyQt6.QtWidgets import QFileDialog, QTableWidgetItem, QWidget, QHBoxLayout, QLineEdit, QProgressBar, QHeaderView, QMessageBox

# To build an .exe file -> python -m PyInstaller ctptool.spec

class MainWindow(QtWidgets.QMainWindow):

    populate_signal = QtCore.pyqtSignal()
    status_signal = QtCore.pyqtSignal(str)

    project_name = "No Project Selected"
    user_name = "Unknown"
    missing = False
    read = True
    theme_change = False
    data_saved = True

    def __init__(self):

        super().__init__()

        self.io_list = []
        self.manual_list = []
        self.sequence_list = []
        self.user_name = self.get_user_name()
        self.settings = QtCore.QSettings("BOSCH", "CTP")
        self.current_theme = self.settings.value("theme", "White")

        self.setWindowIcon(QtGui.QIcon(self.icon_path("icon.ico")))
        self.setWindowTitle("OpCon Classic CTP")
        self.resize(1280, 720)

        self.status_signal.connect(self.update_status_label)

        central_widget = QWidget()

        self.setCentralWidget(central_widget)

        self.project_label = QtWidgets.QLabel(" CURRENT PROJECT :")
        self.selected_project = QtWidgets.QLabel("No Project Selected")

        new_button = QtWidgets.QPushButton("New Project")
        new_button.clicked.connect(self.new_project)
        load_button = QtWidgets.QPushButton("Load Data")
        load_button.clicked.connect(self.load_data)
        save_button = QtWidgets.QPushButton("Save Data")
        save_button.clicked.connect(self.save_data)
        export_button = QtWidgets.QPushButton("Export Data")
        export_button.clicked.connect(self.export_data)

        self.theme_label = QtWidgets.QLabel("     THEME :")
        self.theme_selector = QtWidgets.QComboBox()
        self.theme_selector.addItems(["White", "Gray", "Deep Blue", "Wine", "Emerald", "Dark Gold", "Lilac"])

        index = self.theme_selector.findText(self.current_theme)

        if index != -1:

            self.theme_selector.setCurrentIndex(index)

        else:

            self.theme_selector.setCurrentIndex(0)

        self.theme_selector.currentTextChanged.connect(self.change_theme)

        self.user_name_label = QtWidgets.QLabel("")
        self.user_name_label.setText(f"     USER :  {self.user_name} ")

        top_bar = QtWidgets.QHBoxLayout()
        top_bar.addWidget(self.project_label)
        top_bar.addWidget(self.selected_project)
        top_bar.addStretch()
        top_bar.addWidget(new_button)
        top_bar.addWidget(load_button)
        top_bar.addWidget(save_button)
        top_bar.addWidget(export_button)
        top_bar.addWidget(self.theme_label)
        top_bar.addWidget(self.theme_selector)
        top_bar.addWidget(self.user_name_label)

        self.overview_widget = QWidget()
        self.overview_layout = QtWidgets.QVBoxLayout(self.overview_widget)
        self.overview_layout.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.overview_layout.addStretch()
        self.overview_progress_labels = {}
        self.overview_progress_bars = {}
        self.progress_label_max_width = 0
        self.create_progress_display("IO", self.overview_layout)
        self.create_progress_display("MANUAL", self.overview_layout)
        self.create_progress_display("SEQUENCE", self.overview_layout)
        self.overview_layout.addStretch()

        self.io_table = self.create_table(["Module", "BMK", "Address", "Description", "Check", "Comment"], stretch_cols=[5])
        self.manual_table = self.create_table(["Name", "BAS", "Check", "Comment", "WRK", "Check", "Comment"], stretch_cols=[3, 6])
        self.sequence_table = self.create_table(["Name", "Work Position", "Address", "Check", "Comment"], stretch_cols=[4])

        self.tabs = QtWidgets.QTabWidget()
        self.tabs.addTab(self.overview_widget, "OVERVIEW")
        self.tabs.addTab(self.io_table, "IO")
        self.tabs.addTab(self.manual_table, "MANUAL")
        self.tabs.addTab(self.sequence_table, "SEQUENCE")
        self.populate_signal.connect(self.populate_tables)
        self.tabs.currentChanged.connect(self.on_tab_changed)

        self.progress_label = QtWidgets.QLabel(" PROGRESS WILL SHOWN HERE : ")
        self.progress_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedHeight(15)

        progress_layout = QtWidgets.QHBoxLayout()
        progress_layout.addWidget(self.progress_label)
        progress_layout.addWidget(self.progress_bar)

        self.status_label = QtWidgets.QLabel("INFO: Please create a new project or load your saved data.")
        self.status_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)

        layout = QtWidgets.QVBoxLayout(central_widget)
        layout.addLayout(top_bar)
        layout.addWidget(self.tabs)
        layout.addLayout(progress_layout)
        layout.addWidget(self.status_label)

    @QtCore.pyqtSlot(str)
    def update_status_label(self, message):

        self.status_label.setText(message)

    def icon_path(self, path):

        if hasattr(sys, '_MEIPASS'):

            return os.path.join(sys._MEIPASS, path)

        return os.path.join(os.path.abspath("."), path)

    def get_user_name(self):

        return getpass.getuser()

    def new_project(self):

        def browse_folder():

            folder = QFileDialog.getExistingDirectory(self, "Select Main Project Location")

            if folder:

                self.project_name = os.path.basename(folder)
                self.status_signal.emit("Reading Files...")

                threading.Thread(target=read_files_thread, args=(folder,), daemon=True).start()

            else:

                self.status_label.setText("INFO: New project cancelled.")

        def read_files_thread(folder):

            self.io_list.clear()
            self.manual_list.clear()
            self.sequence_list.clear()

            if not read_files(folder, self.io_list, self.manual_list, self.sequence_list):

                self.missing = True

            if self.read:

                self.status_signal.emit("Processing Files...")

            self.populate_signal.emit()

        def read_files(exp_path, io_list, manual_list, sequence_list):

            file_counter = 0

            for file in os.listdir(exp_path):

                if file.lower().endswith('.exp') and os.path.isfile(os.path.join(exp_path, file)) and file_counter < 3:

                    try:

                        with open(os.path.join(exp_path, file), 'r', encoding='utf-8') as f:

                            if file.startswith('Var'):

                                io_full_content = [line.strip() for line in f.readlines() if line.strip()]
                                process_file("io", io_full_content, io_list)
                                file_counter += 1

                            elif file.startswith('BM'):

                                manual_full_content = [line.strip() for line in f.readlines() if line.strip()]
                                process_file("manual", manual_full_content, manual_list)
                                file_counter += 1

                            elif file.startswith('SfcDiag'):

                                sequence_full_content = [line.strip() for line in f.readlines() if line.strip()]
                                process_file("sequence", sequence_full_content, sequence_list)
                                file_counter += 1

                    except Exception as e:

                        self.read = False
                        self.status_signal.emit(f"ERROR: Couldn't read the files! Please make sure that you navigate the main project folder and show this message to your developer -> {e}")

                        return False

            if file_counter == 3:

                return True

            else:

                return False

        def process_file(content_type, full_content, content_list):

            content_list.clear()

            if content_type == "io":

                current_module = None

                for line in full_content:

                    module_match = re.search(r'\(\*\s*-+\s*([A-Za-z0-9_]+)\s*-+\s*\*\)', line)

                    if module_match:
                        current_module = module_match.group(1)

                        continue

                    bmk_match = re.match(r'^(\S+)', line)
                    bmk = bmk_match.group(1) if bmk_match else "NULL"

                    address_match = re.search(r'AT\s+([^:\s]+)', line)
                    address = address_match.group(1) if address_match else "NULL"

                    description_match = re.search(r'\(\*e\s+(.*?)\s*\*\)', line)
                    description = description_match.group(1) if description_match else "NULL"

                    if bmk == "(*":

                        bmk = "NULL"

                    if current_module and address != "NULL":

                        content_list.append([current_module, bmk, address, description, (".", ""), ("", "")])

                print("IO file was processed successfully.")

            elif content_type == "manual":

                current_title = None
                current_bas = None
                current_wrk = None

                for line in full_content:

                    if "END_PROGRAM" in line:

                        break

                    title_match = re.search(r'\(\*\s*(\d+_[A-Za-z0-9]+_\d+K\d+)\s*\*\)', line)

                    if title_match:

                        if current_title:

                            content_list.append([current_title, current_bas if current_bas else "", (".", ""), ("", ""),
                                                 current_wrk if current_wrk else "", (".", ""), ("", "")])

                        current_title = title_match.group(1)
                        current_bas = None
                        current_wrk = None

                        continue

                    bas_match = re.search(r'\bBAS_\d+K\d+\b', line)

                    if bas_match:

                        current_bas = bas_match.group()

                        continue

                    wrk_match = re.search(r'\bWRK_\d+K\d+\b', line)

                    if wrk_match:

                        current_wrk = wrk_match.group()

                        continue

                if current_title:

                    content_list.append([current_title, current_bas if current_bas else "", (".", ""), ("", ""), current_wrk if current_wrk else "", (".", ""), ("", "")])

                print("Manual file was processed successfully.")

            elif content_type == "sequence":

                i = 0

                while i < len(full_content) - 1:

                    line1 = full_content[i]
                    line1_match = re.search(r'\(\*\s*(.*?)\s*\*\)', line1)

                    line2 = full_content[i + 1]
                    line2_match = re.match(r'(\S+)\s*:=\s*(.+);', line2)

                    if line1_match and line2_match:

                        name = line1_match.group(1)
                        work_position = line2_match.group(1)
                        address = line2_match.group(2)
                        content_list.append([name, work_position, address, (".", ""), ("", "")])
                        i += 2

                    else:

                        i += 1

                print("Sequence file was processed successfully.")

            else:

                print("Invalid content type! Exiting from task...")

        browse_folder()
        self.data_saved = False

    def load_data(self):

        def select_file():

            options = QFileDialog.Option.DontUseNativeDialog
            file_name, _ = QFileDialog.getOpenFileName(self, "Load Data", "", "JSON Files (*.json);;All Files (*)", options=options)

            if not file_name:

                self.status_label.setText("INFO: Load cancelled.")

                return

            self.status_label.setText("Loading data...")

            transfer_data(file_name)

        def transfer_data(file_name):

            import json

            try:

                with open(file_name, "r", encoding="utf-8") as f:

                    data = json.load(f)

                self.project_name = data.get("project_name", "Untitled")
                self.io_list = [[make_tuple(cell, str) if i in [4, 5] else cell for i, cell in enumerate(row)] for row in data.get("io_list", [])]
                self.manual_list = [[make_tuple(cell, str) if i in [2, 3, 5, 6] else cell for i, cell in enumerate(row)] for row in data.get("manual_list", [])]
                self.sequence_list = [[make_tuple(cell, str) if i in [3, 4] else cell for i, cell in enumerate(row)] for row in data.get("sequence_list", [])]

                self.populate_tables()
                self.selected_project.setText(self.project_name)
                self.status_label.setText("INFO: Data loaded successfully.")
                self.data_saved = True

            except Exception as e:

                self.status_label.setText(f"ERROR: Couldn't load the data! Please show this message to your developer -> {e}")

        def make_tuple(value, expected_type):

            if isinstance(value, list) and len(value) == 2 and isinstance(value[0], expected_type):

                return tuple(value)

            return value

        select_file()

    def save_data(self):

        def prepare_data():

            options = QFileDialog.Option.DontUseNativeDialog
            default_dir = self.selected_project.text()
            default_file_name = self.project_name + ".json"
            file_name, _ = QFileDialog.getSaveFileName(self, "Save Data", os.path.join(default_dir, default_file_name), "JSON Files (*.json);;All Files (*)", options=options)

            if not file_name:

                self.status_label.setText("INFO: Save cancelled.")

                return

            base_name = os.path.basename(file_name)
            name_without_ext = os.path.splitext(base_name)[0]
            self.project_name = name_without_ext

            data = {"project_name": self.project_name, "io_list": self.io_list, "manual_list": self.manual_list, "sequence_list": self.sequence_list}
            threading.Thread(target=save_json, args=(file_name, data,), daemon=True).start()

        def save_json(file_name, data):

            import json

            try:

                self.status_signal.emit("Saving data...")

                with open(file_name, "w", encoding="utf-8") as f:

                    json.dump(data, f, indent=4)

                self.status_signal.emit("INFO: Data saved successfully.")

            except Exception as e:

                self.status_signal.emit(f"ERROR: Couldn't save the data! Please show this message to your developer -> {e}")

            self.data_saved = True

        if self.selected_project.text() == "No Project Selected":

            self.status_label.setText("WARNING: Please select a project first.")

            return

        else:

            prepare_data()

    def export_data(self):

        def select_location():

            options = QFileDialog.Option.DontUseNativeDialog
            default_file_name = self.project_name + ".xlsx"
            file_name, _ = QFileDialog.getSaveFileName(self, "Export", os.path.join(self.selected_project.text(), default_file_name), "Excel Files (*.xlsx);;All Files (*)", options=options)

            if not file_name:

                self.status_label.setText("INFO: Export cancelled.")

                return

            else:

                prepare_data(file_name)

        def prepare_data(file_name):

            io_cleaned = clean_data_list(self.io_list)
            io_progress = self.calculate_progress("io")[1]
            manual_cleaned = clean_data_list(self.manual_list)
            manual_progress = self.calculate_progress("manual")[1]
            sequence_cleaned = clean_data_list(self.sequence_list)
            sequence_progress = self.calculate_progress("sequence")[1]
            total_progress = self.calculate_total_progress()

            threading.Thread(target=transfer_data, args=(file_name, io_cleaned, io_progress, manual_cleaned, manual_progress, sequence_cleaned, sequence_progress, total_progress,), daemon=True).start()

        def clean_data_list(data_list):

            cleaned = []

            for row in data_list:

                cleaned_row = []

                for item in row:

                    if isinstance(item, (list, tuple)):

                        if isinstance(item[0], str) and item[0] in ["X", "OK", "N/A"]:

                            if item[0] == "X":

                                cleaned_row.append("X")

                            elif item[0] == "OK":

                                cleaned_row.append("OK")

                            else:

                                cleaned_row.append("N/A")

                        elif isinstance(item[0], str):

                            if item[0].strip() == ".":

                                cleaned_row.append("X")

                            elif item[0].strip() == "":

                                cleaned_row.append("")

                            elif item[0].strip().startswith(("=", "+", "-")):

                                cleaned_row.append(" " + item[0])

                            else:

                                cleaned_row.append(item[0])

                    elif isinstance(item, str):

                        if item.strip().startswith(("=", "+", "-")):

                            cleaned_row.append(" " + item)

                        else:

                            cleaned_row.append(item)

                    else:

                        cleaned_row.append(item)

                cleaned.append(cleaned_row)

            return cleaned

        def transfer_data(file_name, io_cleaned, io_progress, manual_cleaned, manual_progress, sequence_cleaned, sequence_progress, total_progress):

            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment

            try:

                self.status_signal.emit("Exporting data...")

                with pd.ExcelWriter(file_name, engine="openpyxl") as writer:

                    df_io = pd.DataFrame(io_cleaned, columns=["Module", "BMK", "Address", "Description", "Check", "Comment"])
                    df_io.to_excel(writer, sheet_name="IO", index=False)
                    df_manual = pd.DataFrame(manual_cleaned, columns=["Name", "BAS", "Check", "Comment1", "WRK", "Check", "Comment2"])
                    df_manual.to_excel(writer, sheet_name="MANUAL", index=False)
                    df_seq = pd.DataFrame(sequence_cleaned, columns=["Name", "Work Position", "Address", "Check", "Comment"])
                    df_seq.to_excel(writer, sheet_name="SEQUENCE", index=False)

                wb = load_workbook(file_name)

                for sheet_name in ["IO", "MANUAL", "SEQUENCE"]:

                    ws = wb[sheet_name]

                    for col in ws.columns:

                        max_length = 0
                        column = col[0].column_letter

                        for cell in col:

                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                            try:

                                if len(str(cell.value)) > max_length:

                                    max_length = len(str(cell.value)) + 1

                            except:

                                pass

                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column].width = adjusted_width

                ws_io = wb["IO"]
                set_progress_cell(ws_io, 1, 7, f"IO Progress: {io_progress}%")
                ws_total = wb["IO"]
                set_progress_cell(ws_total, 1, 8, f"TOTAL Progress: {total_progress}%")
                ws_manual = wb["MANUAL"]
                set_progress_cell(ws_manual, 1, 8, f"MANUAL Progress: {manual_progress}%")
                ws_seq = wb["SEQUENCE"]
                set_progress_cell(ws_seq, 1, 6, f"SEQUENCE Progress: {sequence_progress}%")

                wb.save(file_name)
                self.status_signal.emit("INFO: Data exported successfully.")

            except Exception as e:

                self.status_signal.emit(f"ERROR: Couldn't export the data! Please show this message to your developer -> {e}")

        def set_progress_cell(ws, row, col, progress_text):

            from openpyxl.styles import Alignment, Border, Font, Side

            cell = ws.cell(row=row, column=col)
            cell.value = progress_text
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = thin_border
            column_letter = cell.column_letter
            ws.column_dimensions[column_letter].width = len(progress_text) + 5
            ws.row_dimensions[row].height = 20

        if not self.io_list and not self.manual_list and not self.sequence_list:

            self.status_label.setText("WARNING: Please select a project first.")

            return

        else:

            select_location()

    def change_theme(self, theme_name):

        def apply_theme():

            app = QtWidgets.QApplication.instance()

            self.theme_change = True

            if theme_name == "White":

                set_white_palette(app)

            elif theme_name == "Gray":

                set_gray_palette(app)

            elif theme_name == "Deep Blue":

                set_deep_blue_palette(app)

            elif theme_name == "Wine":

                set_wine_palette(app)

            elif theme_name == "Emerald":

                set_emerald_palette(app)

            elif theme_name == "Dark Gold":

                set_dark_gold_palette(app)

            elif theme_name == "Lilac":

                set_lilac_palette(app)

            settings = QtCore.QSettings("BOSCH", "CTP")
            settings.setValue("theme", theme_name)

            self.populate_tables()
            self.theme_change = False

        def set_white_palette(app):

            app.setStyle("Fusion")

            palette = QtGui.QPalette()
            palette.setColor(QtGui.QPalette.ColorRole.Window, QtGui.QColor(240, 240, 240))
            palette.setColor(QtGui.QPalette.ColorRole.WindowText, QtCore.Qt.GlobalColor.black)
            palette.setColor(QtGui.QPalette.ColorRole.Base, QtGui.QColor(230, 230, 230))
            palette.setColor(QtGui.QPalette.ColorRole.AlternateBase, QtGui.QColor(245, 245, 245))
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipBase, QtGui.QColor(240, 240, 240))
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipText, QtCore.Qt.GlobalColor.black)
            palette.setColor(QtGui.QPalette.ColorRole.Text, QtCore.Qt.GlobalColor.black)
            palette.setColor(QtGui.QPalette.ColorRole.Button, QtGui.QColor(245, 245, 245))
            palette.setColor(QtGui.QPalette.ColorRole.ButtonText, QtCore.Qt.GlobalColor.black)
            palette.setColor(QtGui.QPalette.ColorRole.BrightText, QtGui.QColor(255, 0, 0))
            palette.setColor(QtGui.QPalette.ColorRole.Link, QtGui.QColor(115, 115, 115, 180))
            palette.setColor(QtGui.QPalette.ColorRole.Highlight, QtGui.QColor(115, 115, 115, 180))
            palette.setColor(QtGui.QPalette.ColorRole.HighlightedText, QtCore.Qt.GlobalColor.white)

            app.setPalette(palette)

        def set_gray_palette(app):

            app.setStyle("Fusion")

            palette = QtGui.QPalette()
            palette.setColor(QtGui.QPalette.ColorRole.Window, QtGui.QColor(53, 53, 53))
            palette.setColor(QtGui.QPalette.ColorRole.WindowText, QtCore.Qt.GlobalColor.white)
            palette.setColor(QtGui.QPalette.ColorRole.Base, QtGui.QColor(43, 43, 43))
            palette.setColor(QtGui.QPalette.ColorRole.AlternateBase, QtGui.QColor(58, 58, 58))
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipBase, QtGui.QColor(53, 53, 53))
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipText, QtCore.Qt.GlobalColor.white)
            palette.setColor(QtGui.QPalette.ColorRole.Text, QtCore.Qt.GlobalColor.white)
            palette.setColor(QtGui.QPalette.ColorRole.Button, QtGui.QColor(53, 53, 53))
            palette.setColor(QtGui.QPalette.ColorRole.ButtonText, QtCore.Qt.GlobalColor.white)
            palette.setColor(QtGui.QPalette.ColorRole.BrightText, QtCore.Qt.GlobalColor.red)
            palette.setColor(QtGui.QPalette.ColorRole.Link, QtGui.QColor(215, 215, 215, 180))
            palette.setColor(QtGui.QPalette.ColorRole.Highlight, QtGui.QColor(215, 215, 215, 180))
            palette.setColor(QtGui.QPalette.ColorRole.HighlightedText, QtCore.Qt.GlobalColor.black)

            app.setPalette(palette)

        def set_deep_blue_palette(app):

            app.setStyle("Fusion")

            palette = QtGui.QPalette()
            palette.setColor(QtGui.QPalette.ColorRole.Window, QtGui.QColor(30, 40, 50))
            palette.setColor(QtGui.QPalette.ColorRole.WindowText, QtCore.Qt.GlobalColor.white)
            palette.setColor(QtGui.QPalette.ColorRole.Base, QtGui.QColor(20, 30, 40))
            palette.setColor(QtGui.QPalette.ColorRole.AlternateBase, QtGui.QColor(35, 45, 55))
            palette.setColor(QtGui.QPalette.ColorRole.Text, QtCore.Qt.GlobalColor.white)
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipBase, QtGui.QColor(30, 40, 50))
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipText, QtCore.Qt.GlobalColor.white)
            palette.setColor(QtGui.QPalette.ColorRole.Button, QtGui.QColor(40, 50, 60))
            palette.setColor(QtGui.QPalette.ColorRole.ButtonText, QtCore.Qt.GlobalColor.white)
            palette.setColor(QtGui.QPalette.ColorRole.BrightText, QtGui.QColor(255, 100, 100))
            palette.setColor(QtGui.QPalette.ColorRole.Link, QtGui.QColor(120, 180, 240, 180))
            palette.setColor(QtGui.QPalette.ColorRole.Highlight, QtGui.QColor(120, 180, 240, 180))
            palette.setColor(QtGui.QPalette.ColorRole.HighlightedText, QtCore.Qt.GlobalColor.black)

            app.setPalette(palette)

        def set_wine_palette(app):

            app.setStyle("Fusion")

            palette = QtGui.QPalette()
            palette.setColor(QtGui.QPalette.ColorRole.Window, QtGui.QColor(45, 10, 20))
            palette.setColor(QtGui.QPalette.ColorRole.WindowText, QtGui.QColor(240, 230, 200))
            palette.setColor(QtGui.QPalette.ColorRole.Base, QtGui.QColor(35, 0, 10))
            palette.setColor(QtGui.QPalette.ColorRole.AlternateBase, QtGui.QColor(50, 15, 25))
            palette.setColor(QtGui.QPalette.ColorRole.Text, QtGui.QColor(240, 230, 200))
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipBase, QtGui.QColor(45, 10, 20))
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipText, QtGui.QColor(240, 230, 200))
            palette.setColor(QtGui.QPalette.ColorRole.Button, QtGui.QColor(60, 20, 30))
            palette.setColor(QtGui.QPalette.ColorRole.ButtonText, QtGui.QColor(240, 230, 200))
            palette.setColor(QtGui.QPalette.ColorRole.BrightText, QtGui.QColor(255, 150, 150))
            palette.setColor(QtGui.QPalette.ColorRole.Link, QtGui.QColor(210, 0, 60, 180))
            palette.setColor(QtGui.QPalette.ColorRole.Highlight, QtGui.QColor(210, 0, 60, 180))
            palette.setColor(QtGui.QPalette.ColorRole.HighlightedText, QtCore.Qt.GlobalColor.black)

            app.setPalette(palette)

        def set_emerald_palette(app):

            app.setStyle("Fusion")

            palette = QtGui.QPalette()
            palette.setColor(QtGui.QPalette.ColorRole.Window, QtGui.QColor(25, 60, 45))
            palette.setColor(QtGui.QPalette.ColorRole.WindowText, QtGui.QColor(240, 230, 200))
            palette.setColor(QtGui.QPalette.ColorRole.Base, QtGui.QColor(15, 50, 35))
            palette.setColor(QtGui.QPalette.ColorRole.AlternateBase, QtGui.QColor(30, 65, 50))
            palette.setColor(QtGui.QPalette.ColorRole.Text, QtGui.QColor(240, 230, 200))
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipBase, QtGui.QColor(25, 60, 45))
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipText, QtGui.QColor(240, 230, 200))
            palette.setColor(QtGui.QPalette.ColorRole.Button, QtGui.QColor(40, 80, 65))
            palette.setColor(QtGui.QPalette.ColorRole.ButtonText, QtGui.QColor(240, 230, 200))
            palette.setColor(QtGui.QPalette.ColorRole.BrightText, QtGui.QColor(255, 120, 120))
            palette.setColor(QtGui.QPalette.ColorRole.Link, QtGui.QColor(75, 250, 175, 180))
            palette.setColor(QtGui.QPalette.ColorRole.Highlight, QtGui.QColor(75, 250, 175, 180))
            palette.setColor(QtGui.QPalette.ColorRole.HighlightedText, QtCore.Qt.GlobalColor.black)

            app.setPalette(palette)

        def set_dark_gold_palette(app):

            app.setStyle("Fusion")

            palette = QtGui.QPalette()
            palette.setColor(QtGui.QPalette.ColorRole.Window, QtGui.QColor(50, 40, 20))
            palette.setColor(QtGui.QPalette.ColorRole.WindowText, QtGui.QColor(255, 215, 200))
            palette.setColor(QtGui.QPalette.ColorRole.Base, QtGui.QColor(40, 30, 15))
            palette.setColor(QtGui.QPalette.ColorRole.AlternateBase, QtGui.QColor(55, 45, 25))
            palette.setColor(QtGui.QPalette.ColorRole.Text, QtGui.QColor(255, 220, 200))
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipBase, QtGui.QColor(50, 40, 20))
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipText, QtGui.QColor(255, 235, 200))
            palette.setColor(QtGui.QPalette.ColorRole.Button, QtGui.QColor(65, 60, 40))
            palette.setColor(QtGui.QPalette.ColorRole.ButtonText, QtGui.QColor(255, 215, 200))
            palette.setColor(QtGui.QPalette.ColorRole.BrightText, QtGui.QColor(255, 255, 200))
            palette.setColor(QtGui.QPalette.ColorRole.Link, QtGui.QColor(240, 180, 90, 180))
            palette.setColor(QtGui.QPalette.ColorRole.Highlight, QtGui.QColor(240, 180, 90, 180))
            palette.setColor(QtGui.QPalette.ColorRole.HighlightedText, QtCore.Qt.GlobalColor.black)

            app.setPalette(palette)

        def set_lilac_palette(app):

            app.setStyle("Fusion")

            palette = QtGui.QPalette()
            palette.setColor(QtGui.QPalette.ColorRole.Window, QtGui.QColor(60, 40, 60))
            palette.setColor(QtGui.QPalette.ColorRole.WindowText, QtGui.QColor(255, 200, 230))
            palette.setColor(QtGui.QPalette.ColorRole.Base, QtGui.QColor(50, 30, 50))
            palette.setColor(QtGui.QPalette.ColorRole.AlternateBase, QtGui.QColor(65, 45, 65))
            palette.setColor(QtGui.QPalette.ColorRole.Text, QtGui.QColor(255, 210, 235))
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipBase, QtGui.QColor(60, 40, 60))
            palette.setColor(QtGui.QPalette.ColorRole.ToolTipText, QtGui.QColor(255, 220, 240))
            palette.setColor(QtGui.QPalette.ColorRole.Button, QtGui.QColor(80, 60, 80))
            palette.setColor(QtGui.QPalette.ColorRole.ButtonText, QtGui.QColor(255, 200, 230))
            palette.setColor(QtGui.QPalette.ColorRole.BrightText, QtGui.QColor(255, 255, 255))
            palette.setColor(QtGui.QPalette.ColorRole.Link, QtGui.QColor(250, 150, 250, 180))
            palette.setColor(QtGui.QPalette.ColorRole.Highlight, QtGui.QColor(250, 150, 250, 180))
            palette.setColor(QtGui.QPalette.ColorRole.HighlightedText, QtCore.Qt.GlobalColor.black)

            app.setPalette(palette)

        apply_theme()

    def get_alternate_row_colors(self):

        palette = QtWidgets.QApplication.instance().palette()
        base_color = palette.color(QtGui.QPalette.ColorRole.Base)
        alternate_base_color = palette.color(QtGui.QPalette.ColorRole.AlternateBase)

        return base_color, alternate_base_color

    def create_table(self, headers, stretch_cols=[]):

        table = QtWidgets.QTableWidget()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.horizontalHeader().setStretchLastSection(True)
        table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        header = table.horizontalHeader()

        for i in range(len(headers)):

            if i in stretch_cols:

                header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)

            else:

                header.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)

        table.verticalHeader().setDefaultSectionSize(40)
        table.verticalHeader().setMinimumWidth(32)
        table.verticalHeader().setMaximumWidth(32)
        table.setAlternatingRowColors(True)

        return table

    @QtCore.pyqtSlot()
    def populate_tables(self):

        def populate_table(table, data_list, special_cols, tab_name):

            table.setRowCount(len(data_list))
            base_color, alternate_base_color = self.get_alternate_row_colors()

            for row_idx, row_data in enumerate(data_list):

                table.setRowHeight(row_idx, 40)

                for col_idx, value in enumerate(row_data):

                    if col_idx in special_cols:

                        if (tab_name == "io" and col_idx == 4) or (tab_name == "manual" and (col_idx == 2 or col_idx == 5)) or (tab_name == "sequence" and col_idx == 3):

                            val_str, user = value
                            widget = QWidget()
                            layout = QHBoxLayout()
                            layout.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                            cb = QtWidgets.QComboBox()
                            cb.addItems(["X", "OK", "N/A"])
                            index = cb.findText(val_str)

                            if index != -1:

                                cb.setCurrentIndex(index)

                            else:

                                cb.setCurrentIndex(0)

                            cb.setToolTip(f"Last modified by: {user}" if user else "No modifications yet.")
                            cb.currentTextChanged.connect(lambda text, r=row_idx, c=col_idx, t=tab_name: self.combobox_changed(t, r, c, text))
                            layout.addWidget(cb)
                            widget.setLayout(layout)
                            table.setCellWidget(row_idx, col_idx, widget)

                        elif isinstance(value, tuple) and isinstance(value[0], str):

                            text, user = value
                            le = QLineEdit()
                            le.setText(text)
                            le.setToolTip(f"Last modified by: {user}" if user else "No modifications yet.")
                            le.textChanged.connect(lambda text, r=row_idx, c=col_idx, t=tab_name: self.comment_changed(t, r, c, text))

                            if row_idx % 2 == 0:

                                le.setStyleSheet(f"QLineEdit {{ background-color: {base_color.name()}; }}")

                            else:

                                le.setStyleSheet(f"QLineEdit {{ background-color: {alternate_base_color.name()}; }}")

                            table.setCellWidget(row_idx, col_idx, le)

                        else:

                            le = QLineEdit()
                            le.setText(value)
                            le.textChanged.connect(lambda text, r=row_idx, c=col_idx, t=tab_name: self.comment_changed(t, r, c, text))

                            if row_idx % 2 == 0:

                                le.setStyleSheet(f"QLineEdit {{ background-color: {base_color.name()}; }}")

                            else:

                                le.setStyleSheet(f"QLineEdit {{ background-color: {alternate_base_color.name()}; }}")

                            table.setCellWidget(row_idx, col_idx, le)

                    else:

                        item = QTableWidgetItem(str(value))
                        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                        table.setItem(row_idx, col_idx, item)

            self.update_progress_bar(tab_name)

        tab_name = self.tabs.tabText(self.tabs.currentIndex()).lower()

        populate_table(self.io_table, self.io_list, [4, 5], "io")
        populate_table(self.manual_table, self.manual_list, [2, 3, 5, 6], "manual")
        populate_table(self.sequence_table, self.sequence_list, [3, 4], "sequence")

        if self.read:

            if self.missing:

                self.status_label.setText("WARNING: Some files were not found. Some tabs will be empty. Please try again and make sure that you have selected right location to use this program efficiently.")

            elif not self.missing and not self.theme_change:

                self.status_label.setText("INFO: New project is created successfully.")

        self.selected_project.setText(self.project_name)

        if not tab_name == -1:

            self.update_progress_bar(tab_name)

        self.update_overview_progress()

    def combobox_changed(self, tab, row, col, text):

        value = (text, self.user_name)

        if tab == "io":

            self.io_list[row][col] = value
            table = self.io_table

        elif tab == "manual":

            self.manual_list[row][col] = value
            table = self.manual_table

        elif tab == "sequence":

            self.sequence_list[row][col] = value
            table = self.sequence_table

        if table:

            cell_widget = table.cellWidget(row, col)

            if cell_widget:

                combobox = cell_widget.findChild(QtWidgets.QComboBox)

                if combobox:
                    combobox.setToolTip(f"Last modified by: {self.user_name}")

        self.update_progress_bar(tab)
        self.update_overview_progress()
        self.data_saved = False

    def comment_changed(self, tab, row, col, text):

        value = (text, self.user_name)

        if tab == "io":

            self.io_list[row][col] = value
            table = self.io_table

        elif tab == "manual":

            self.manual_list[row][col] = value
            table = self.manual_table

        elif tab == "sequence":

            self.sequence_list[row][col] = value
            table = self.sequence_table

        if table:

            line_edit = table.cellWidget(row, col)

            if line_edit:
                line_edit.setToolTip(f"Last modified by: {self.user_name}")

        self.data_saved = False

    def on_tab_changed(self, index):

        tab_name = self.tabs.tabText(index).lower()

        self.update_progress_bar(tab_name)

    def create_progress_display(self, tab_name, parent_layout):

        def update_progress_label_widths():

            for label in self.overview_progress_labels.values():

                label.setFixedWidth(self.progress_label_max_width)

        individual_progress_widget = QWidget()
        individual_layout = QHBoxLayout(individual_progress_widget)
        individual_layout.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        individual_layout.setSpacing(20)
        individual_layout.setContentsMargins(0, 10, 0, 10)

        label = QtWidgets.QLabel(f"{tab_name} PROGRESS :")
        label.setAlignment(QtCore.Qt.AlignmentFlag.AlignRight | QtCore.Qt.AlignmentFlag.AlignVCenter)

        font_metrics = QtGui.QFontMetrics(label.font())
        text_width = font_metrics.horizontalAdvance(label.text()) + 10

        if text_width > self.progress_label_max_width:

            self.progress_label_max_width = text_width

        bar = QProgressBar()
        bar.setMinimumHeight(20)
        bar.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        bar.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Fixed)

        individual_layout.addStretch()
        individual_layout.addWidget(label)
        individual_layout.addWidget(bar)
        individual_layout.addStretch()
        parent_layout.addWidget(individual_progress_widget)

        self.overview_progress_labels[tab_name.lower()] = label
        self.overview_progress_bars[tab_name.lower()] = bar

        update_progress_label_widths()

    def calculate_progress(self, tab):

        value = []

        if tab == "io":

            total = len(self.io_list)
            total -= sum((isinstance(r[4], tuple) and r[4][0] == "N/A") for r in self.io_list)
            checked = sum(1 for r in self.io_list if isinstance(r[4], tuple) and r[4][0] == "OK")
            value.append("IO")

        elif tab == "manual":

            total = 2 * len(self.manual_list)
            total -= sum((isinstance(r[2], tuple) and r[2][0] == "N/A") + (isinstance(r[5], tuple) and r[5][0] == "N/A") for r in self.manual_list)
            checked = sum((isinstance(r[2], tuple) and r[2][0] == "OK") + (isinstance(r[5], tuple) and r[5][0] == "OK") for r in self.manual_list)
            value.append("MANUAL")

        elif tab == "sequence":

            total = len(self.sequence_list)
            total -= sum((isinstance(r[3], tuple) and r[3][0] == "N/A") for r in self.sequence_list)
            checked = sum(1 for r in self.sequence_list if isinstance(r[3], tuple) and r[3][0] == "OK")
            value.append("SEQUENCE")

        else:

            value.append("")

            return 0

        percent = (checked / total) * 100 if total else 0
        value.append(int(percent))

        return value

    def calculate_total_progress(self):

        total = len(self.io_list)
        total -= sum(1 for r in self.io_list if isinstance(r[4], tuple) and r[4][0] == "N/A")
        total += 2 * len(self.manual_list)
        total -= sum((isinstance(r[2], tuple) and r[2][0] == "N/A") + (isinstance(r[5], tuple) and r[5][0] == "N/A") for r in self.manual_list)
        total += len(self.sequence_list)
        total -= sum(1 for r in self.sequence_list if isinstance(r[3], tuple) and r[3][0] == "N/A")

        checked = sum(1 for r in self.io_list if isinstance(r[4], tuple) and r[4][0] == "OK")
        checked += sum((isinstance(r[2], tuple) and r[2][0] == "OK") + (isinstance(r[5], tuple) and r[5][0] == "OK") for r in self.manual_list)
        checked += sum(1 for r in self.sequence_list if isinstance(r[3], tuple) and r[3][0] == "OK")

        percent = (checked / total) * 100 if total else 0

        return int(percent)

    def update_progress_bar(self, tab):

        if tab == "overview":

            progress_value = self.calculate_total_progress()

            self.progress_label.setText("TOTAL PROGRESS : ")
            self.progress_bar.setValue(progress_value)

        else:
            value = self.calculate_progress(tab)

            self.progress_label.setText(f"{value[0]} PROGRESS : ")
            self.progress_bar.setValue(value[1])

    def update_overview_progress(self):

        progress_data = {"io": self.calculate_progress("io")[1], "manual": self.calculate_progress("manual")[1], "sequence": self.calculate_progress("sequence")[1]}

        for tab, value in progress_data.items():

            if tab in self.overview_progress_bars:

                self.overview_progress_bars[tab].setValue(value)
                self.overview_progress_labels[tab].setText(f"{tab.upper()} PROGRESS :")

        if self.tabs.tabText(self.tabs.currentIndex()).lower() == "overview":

            self.update_progress_bar("overview")

    def closeEvent(self, event):

        if not self.data_saved:

            reply = QMessageBox.question(self, 'Save Before Exit?', "WARNING: Do you want to save your unsaved changes before exit?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel)

            if reply == QMessageBox.StandardButton.Yes:

                self.save_data()

                if self.data_saved:

                    event.accept()

                else:

                    event.ignore()

            elif reply == QMessageBox.StandardButton.No:

                event.accept()

            else:

                event.ignore()

        else:

            event.accept()

    def resizeEvent(self, event):

        def equalize_manual_comments():

            if hasattr(self, 'manual_table'):

                table = self.manual_table
                header = table.horizontalHeader()
                comment1_col = 3
                comment2_col = 6
                total_width = table.viewport().width()
                other_width = 0

                for i in range(table.columnCount()):

                    if i not in [comment1_col, comment2_col]:

                        other_width += header.sectionSize(i)

                remaining_width = max(0, total_width - other_width)
                equal_width = remaining_width // 2
                header.resizeSection(comment1_col, equal_width)
                header.resizeSection(comment2_col, equal_width)

        super().resizeEvent(event)

        equalize_manual_comments()

def main():

    app = QtWidgets.QApplication(sys.argv)
    settings = QtCore.QSettings("BOSCH", "CTP")
    window = MainWindow()
    theme = settings.value("theme", "White")
    window.change_theme(theme)
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()